import pandas as pd
import numpy as np
from pandas.api.types import CategoricalDtype
import openpyxl
from monthreport.rank_by_city import RankByCity

file_out = 'C:/Users/shelly/Desktop/周报数据源/xx周.xlsx'
file_1X_busycell = 'C:/Users/shelly/Desktop/周报数据源/1x忙小区.xls'
# 自己新建的
file_lte_prb_busycell = 'C:/Users/shelly/Desktop/周报数据源/lte prb 忙小区.xlsx'
file_lte_flow = 'C:/Users/shelly/Desktop/周报数据源/LTE关键指标统计报表流量.xlsx'
file_800_flow = 'C:/Users/shelly/Desktop/周报数据源/800M关键指标统计报表流量.xlsx'
file_lte_perform = 'C:/Users/shelly/Desktop/周报数据源/LTE关键指标统计报表性能.xlsx'
file_1x_video_loss = 'C:/Users/shelly/Desktop/周报数据源/1x掉话率.xls'
file_volte_video_loss = 'C:/Users/shelly/Desktop/周报数据源/VOLTE掉话率.xlsx'
file_5G = 'C:/Users/shelly/Desktop/周报数据源/数据共享列表（11月26日新模板）-浙江省20210302.xlsx'
file_5G_Unicom = 'C:/Users/shelly/Desktop/周报数据源/附表1：数据共享列表增加SA_浙江联通0303.xlsx'
file_1X_Erlang = 'C:/Users/shelly/Desktop/周报数据源/1x话务量.xls'
# 需要另存为xls
file_rsrp = 'C:/Users/shelly/Desktop/周报数据源/MR地市级覆盖rsrp2.xls'

# 获取周报
week_wb = openpyxl.load_workbook(file_out)
week_ws = week_wb['汇总']

# 排序标准序列
custom_order = CategoricalDtype(['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省'], ordered=True)
#
city = ['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省']

# 0, prb忙小区
busy_prb = pd.read_excel(file_lte_prb_busycell, sheet_name='prb忙小区')
bc_prb = busy_prb[busy_prb['prb超忙天数'] > 2].loc[:, 'city'].value_counts()
value_bc_prb = [0] * 12
df_bc_prb = pd.DataFrame({'地市': city, '一周PRB忙小区数': value_bc_prb})
# 遍历series ，并设置df_bc_prb实际值
for i, v in bc_prb.items():
    df_bc_prb.loc[df_bc_prb.loc[df_bc_prb['地市'].isin([i])].index.tolist()[0], '一周PRB忙小区数'] = v
    df_bc_prb.loc[11, '一周PRB忙小区数'] += v
# 写入周报
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=157, max_row=168):
    for cell in col:
        value = df_bc_prb.at[cell.row - 157, '一周PRB忙小区数']
        cell.value = value
week_wb.save(file_out)
# 1, 1X忙小区
busycell_1x = pd.read_excel(file_1X_busycell)
bc_1x = busycell_1x[busycell_1x['小区超忙天数'] > 2].loc[:, '地市'].value_counts()  # series 类型
# df赋初始值0
value_bc_1x = [0] * 12
df_bc_1x = pd.DataFrame({'地市': city, '1x忙小区数': value_bc_1x})
# 遍历series ，并设置df_bc_1x实际值
for i, v in bc_1x.items():
    # print(i, v)
    # series中存在的地市，更新value
    df_bc_1x.loc[df_bc_1x.loc[df_bc_1x['地市'].isin([i])].index.tolist()[0], '1x忙小区数'] = v
    # 全省为df各个值得相加 全省的索引是11
    df_bc_1x.loc[11, '1x忙小区数'] += v
# print(df_bc_1x)
# 写入周报
for col in week_ws.iter_cols(min_col=5, max_col=5, min_row=3, max_row=14):
    for cell in col:
        value = df_bc_1x.at[cell.row - 3, '1x忙小区数']
        # 设置格式为数值，保留两位小数
        # cell.number_format = '0.00'
        cell.value = value
week_wb.save(file_out)

# 2, lte 忙小区 区分频段
busy_lte = pd.read_excel(file_lte_prb_busycell, sheet_name='lte忙小区')
# 取得各个频段的忙小区数
busy_800 = busy_lte[(busy_lte['band'] == 'FDD800') & (busy_lte['count'] > 2)].loc[:, 'city'].value_counts()
busy_1800 = busy_lte[(busy_lte['band'] == 'FDD1800') & (busy_lte['count'] > 2)].loc[:, 'city'].value_counts()
busy_2100 = busy_lte[(busy_lte['band'] == 'FDD2100') & (busy_lte['count'] > 2)].loc[:, 'city'].value_counts()

# df 赋初值0
value_800_busy = [0] * 12
value_1800_busy = [0] * 12
value_2100_busy = [0] * 12
df_lte_busy = pd.DataFrame(
    {'地市': city, '800 busy': value_800_busy, '1800 busy': value_1800_busy, '2100 busy': value_2100_busy})

# 用series给df更新值 series中已有地市和全省值
for i, v in busy_800.items():
    df_lte_busy.loc[df_lte_busy.loc[df_lte_busy['地市'].isin([i])].index.tolist()[0], '800 busy'] = v
    df_lte_busy.loc[11, '800 busy'] += v
for i, v in busy_1800.items():
    df_lte_busy.loc[df_lte_busy.loc[df_lte_busy['地市'].isin([i])].index.tolist()[0], '1800 busy'] = v
    df_lte_busy.loc[11, '1800 busy'] += v
for i, v in busy_2100.items():
    df_lte_busy.loc[df_lte_busy.loc[df_lte_busy['地市'].isin([i])].index.tolist()[0], '2100 busy'] = v
    df_lte_busy.loc[11, '2100 busy'] += v
# print(df_lte_busy)

# 写入周报
for col in week_ws.iter_cols(min_col=5, max_col=5, min_row=20, max_row=31):
    for cell in col:
        value = df_lte_busy.at[cell.row - 20, '1800 busy'] + df_lte_busy.at[cell.row - 20, '2100 busy']
        # 设置格式为数值，保留两位小数
        # cell.number_format = '0.00'
        cell.value = value
for col in week_ws.iter_cols(min_col=5, max_col=5, min_row=36, max_row=47):
    for cell in col:
        value = df_lte_busy.at[cell.row - 36, '800 busy']
        # 设置格式为数值，保留两位小数
        # cell.number_format = '0.00'
        cell.value = value
week_wb.save(file_out)

# 3, 4G流量
# 4G总流量
# 读取文件
flow_4G_wb = openpyxl.load_workbook(file_lte_flow)
flow_4G_ws = flow_4G_wb['sheet1']

# 构造df并赋值
city_4G_flow = []
ul_flow = []
dl_flow = []
for col in flow_4G_ws.iter_cols(min_col=2, max_col=2, min_row=5):
    for cell in col:
        city_4G_flow.append(cell.value)

for col in flow_4G_ws.iter_cols(min_col=6, max_col=6, min_row=5):
    for cell in col:
        ul_flow.append(cell.value)

for col in flow_4G_ws.iter_cols(min_col=7, max_col=7, min_row=5):
    for cell in col:
        dl_flow.append(cell.value)

# print(dl_flow)
df_4G_flow = pd.DataFrame({'city': city_4G_flow, 'ul flow': ul_flow, 'dl flow': dl_flow})
# print(df_4G_flow)
df_4G_flow['total 4G flow'] = df_4G_flow['ul flow'] + df_4G_flow['dl flow']
# print(df_4G_flow)
df_4G_flow_avg = pd.DataFrame(df_4G_flow.groupby('city')['total 4G flow'].mean() / 1024 / 1024)
# print(df_4G_flow_avg)
# '浙江'都换作'全省'。 先得到元素所在坐标，再赋值
df_4G_flow_avg.rename({'浙江': '全省'}, inplace=True)
# print(df_4G_flow_avg.index)
# 排序
df_4G_flow_avg.index = df_4G_flow_avg.index.astype(custom_order)
df_4G_flow_avg = df_4G_flow_avg.sort_index()
# print(df_4G_flow_avg)

# 800M流量
flow_800_wb = openpyxl.load_workbook(file_800_flow)
flow_800_ws = flow_800_wb['sheet1']

# 构造df并赋值
city_800_flow = []
ul_800_flow = []
dl_800_flow = []
for col in flow_800_ws.iter_cols(min_col=2, max_col=2, min_row=5):
    for cell in col:
        city_800_flow.append(cell.value)

for col in flow_800_ws.iter_cols(min_col=6, max_col=6, min_row=5):
    for cell in col:
        ul_800_flow.append(cell.value)

for col in flow_800_ws.iter_cols(min_col=7, max_col=7, min_row=5):
    for cell in col:
        dl_800_flow.append(cell.value)
df_800_flow = pd.DataFrame({'city': city_800_flow, 'ul_800_flow': ul_800_flow, 'dl_800_flow': dl_800_flow})
df_800_flow['total_800_flow'] = df_800_flow['ul_800_flow'] + df_800_flow['dl_800_flow']
df_800_flow_avg = pd.DataFrame(df_800_flow.groupby('city')['total_800_flow'].mean() / 1024 / 1024)
df_800_flow_avg.rename({'浙江': '全省'}, inplace=True)
df_800_flow_avg.index = df_800_flow_avg.index.astype(custom_order)
df_800_flow_avg = df_800_flow_avg.sort_index()
# print(df_800_flow_avg)

# 1.8 & 2.1 flow = total 4G flow - total_800_flow
df_4G_flow_avg['1.8&2.1_flow'] = df_4G_flow_avg['total 4G flow'] - df_800_flow_avg['total_800_flow']
df_4G_flow_avg['800_flow'] = df_800_flow_avg['total_800_flow']
# print(df_4G_flow_avg)

# 写入周报
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=20, max_row=31):
    for cell in col:
        value = df_4G_flow_avg.at[cell.row - 20, '1.8&2.1_flow']
        cell.number_format = '0.00'
        cell.value = value
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=36, max_row=47):
    for cell in col:
        value = df_4G_flow_avg.at[cell.row - 36, '800_flow']
        cell.number_format = '0.00'
        cell.value = value
week_wb.save(file_out)

# 4, lte性能指标
perform_lte_wb = openpyxl.load_workbook(file_lte_perform)
perform_lte_ws = perform_lte_wb['sheet1']

df_1x_video_loss = pd.read_excel(file_1x_video_loss)

volte_video_loss_wb = openpyxl.load_workbook(file_volte_video_loss)
volte_video_loss_ws = volte_video_loss_wb['sheet1']

# 构造df
city_lte_perform = []
system_switch_6 = []
RRC_rebuild_7 = []
E_RAB_loss_8 = []
CQI_9 = []

for col in perform_lte_ws.iter_cols(min_col=2, max_col=2, min_row=5, max_row=16):
    for cell in col:
        city_lte_perform.append(cell.value)
for col in perform_lte_ws.iter_cols(min_col=6, max_col=6, min_row=5, max_row=16):
    for cell in col:
        system_switch_6.append(cell.value)
for col in perform_lte_ws.iter_cols(min_col=7, max_col=7, min_row=5, max_row=16):
    for cell in col:
        RRC_rebuild_7.append(cell.value)
for col in perform_lte_ws.iter_cols(min_col=8, max_col=8, min_row=5, max_row=16):
    for cell in col:
        E_RAB_loss_8.append(cell.value)
for col in perform_lte_ws.iter_cols(min_col=9, max_col=9, min_row=5, max_row=16):
    for cell in col:
        CQI_9.append(cell.value)
df_lte_perform = pd.DataFrame(
    {'city': city_lte_perform, 'system_switch_6': system_switch_6, 'RRC_rebuild_7': RRC_rebuild_7,
     'E_RAB_loss_8': E_RAB_loss_8, 'CQI_9': CQI_9})
a = df_lte_perform.loc[df_lte_perform['city'].isin(['浙江'])].index.tolist()
df_lte_perform.loc[a[0], 'city'] = '全省'
df_lte_perform.index = df_lte_perform.iloc[:, 0]
df_lte_perform.index = df_lte_perform.index.astype(custom_order)
df_lte_perform = df_lte_perform.sort_index()
# print(df_lte_perform)

city_VOLTE_2 = []
VOLTE_video_loss_5 = []

for col in volte_video_loss_ws.iter_cols(min_col=2, max_col=2, min_row=4, max_row=15):
    for cell in col:
        city_VOLTE_2.append(cell.value)
for col in volte_video_loss_ws.iter_cols(min_col=5, max_col=5, min_row=4, max_row=15):
    for cell in col:
        VOLTE_video_loss_5.append(cell.value)
df_VOLTE_loss = pd.DataFrame({'city': city_VOLTE_2, 'VOLTE_video_loss_5': VOLTE_video_loss_5})
a = df_VOLTE_loss.loc[df_VOLTE_loss['city'].isin(['浙江'])].index.tolist()
df_VOLTE_loss.loc[a[0], 'city'] = '全省'
df_VOLTE_loss.index = df_VOLTE_loss.iloc[:, 0]
df_VOLTE_loss.index = df_VOLTE_loss.index.astype(custom_order)
df_VOLTE_loss = df_VOLTE_loss.sort_index()
# print(df_VOLTE_loss)

# 处理df
# 1）
# 2）df_1x_video_loss
df_1x_video_loss.index = df_1x_video_loss['城市']
df_1x_video_loss.rename({'浙江': '全省'}, inplace=True)
df_1x_video_loss.index = df_1x_video_loss.index.astype(custom_order)
df_1x_video_loss = df_1x_video_loss.sort_index()
# print(df_1x_video_loss)

# 写入周报

# 1）df_lte_perform
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=90, max_row=101):
    for cell in col:
        value = df_lte_perform.at[cell.row - 90, 'CQI_9']
        cell.number_format = '0.00%'
        cell.value = value/100
for col in week_ws.iter_cols(min_col=3, max_col=3, min_row=90, max_row=101):
    for cell in col:
        value = df_lte_perform.at[cell.row - 90, 'RRC_rebuild_7']
        cell.number_format = '0.000%'
        cell.value = value/100
for col in week_ws.iter_cols(min_col=4, max_col=4, min_row=90, max_row=101):
    for cell in col:
        value = df_lte_perform.at[cell.row - 90, 'system_switch_6']
        cell.number_format = '0.00%'
        cell.value = value/100
for col in week_ws.iter_cols(min_col=5, max_col=5, min_row=90, max_row=101):
    for cell in col:
        value = df_lte_perform.at[cell.row - 90, 'E_RAB_loss_8']
        cell.number_format = '0.000%'
        cell.value = value/100
week_wb.save(file_out)

# 2）df_1x_video_loss
for col in week_ws.iter_cols(min_col=6, max_col=6,min_row=90, max_row=101):
    for cell in col:
        value = df_1x_video_loss.at[cell.row - 90, 'C6.4业务信道掉话率(%)']
        cell.number_format = '0.000%'
        cell.value = value/100
week_wb.save(file_out)

# 3) df_VOLTE_loss
for col in week_ws.iter_cols(min_col=7, max_col=7, min_row=90, max_row=101):
    for cell in col:
        value = df_VOLTE_loss.at[cell.row - 90, 'VOLTE_video_loss_5']
        cell.number_format = '0.00%'
        cell.value = value/100
week_wb.save(file_out)

# 5,5G性能指标 和 5G规模
wb_5G = openpyxl.load_workbook(file_5G)
ws_5G = wb_5G['资源类和PM性能统计类-地市-NSA+SA双模']

# 构造df
city_5G_4 = []
makers_6 = []
CQI_5G_118 = []
SgNB_add_127 = []
SgNB_change_128 = []
gNB_cite_total_8 = []
gNB_cite_share_to_Unicom_12 = []

for col in ws_5G.iter_cols(min_col=4, max_col=4, min_row=6, max_row=33):
    for cell in col:
        city_5G_4.append(cell.value)
for col in ws_5G.iter_cols(min_col=6, max_col=6, min_row=6, max_row=33):
    for cell in col:
        makers_6.append(cell.value)
for col in ws_5G.iter_cols(min_col=118, max_col=118, min_row=6, max_row=33):
    for cell in col:
        CQI_5G_118.append(cell.value)
for col in ws_5G.iter_cols(min_col=127, max_col=127, min_row=6, max_row=33):
    for cell in col:
        SgNB_add_127.append(cell.value)
for col in ws_5G.iter_cols(min_col=128, max_col=128, min_row=6, max_row=33):
    for cell in col:
        SgNB_change_128.append(cell.value)
for col in ws_5G.iter_cols(min_col=8, max_col=8, min_row=6, max_row=33):
    for cell in col:
        gNB_cite_total_8.append(cell.value)
for col in ws_5G.iter_cols(min_col=12, max_col=12, min_row=6, max_row=33):
    for cell in col:
        gNB_cite_share_to_Unicom_12.append(cell.value)

# 联通共享给电信
wb_5G_Unicom = openpyxl.load_workbook(file_5G_Unicom)
ws_5G_Unicom = wb_5G_Unicom['资源类和PM性能统计类-地市(NSA)']

# 构造df
city_5G_Unicom_4 = []
maker_Unicom_6 = []
gNB_cite_from_Unicom_12 = []

for col in ws_5G_Unicom.iter_cols(min_col=4, max_col=4, min_row=6, max_row=36):
    for cell in col:
        city_5G_Unicom_4.append(cell.value)
for col in ws_5G_Unicom.iter_cols(min_col=6, max_col=6, min_row=6, max_row=36):
    for cell in col:
        maker_Unicom_6.append(cell.value)
for col in ws_5G_Unicom.iter_cols(min_col=12, max_col=12, min_row=6, max_row=36):
    for cell in col:
        gNB_cite_from_Unicom_12.append(cell.value)

df_5G = pd.DataFrame({'city': city_5G_4, 'makers_6': makers_6, 'CQI_5G_118': CQI_5G_118, 'SgNB_add_127': SgNB_add_127,
                      'SgNB_change_128': SgNB_change_128, 'gNB_cite_total_8': gNB_cite_total_8,
                      'gNB_cite_share_to_Unicom_12': gNB_cite_share_to_Unicom_12})
df_5G_Unicom = pd.DataFrame({'city': city_5G_Unicom_4, 'maker_Unicom_6': maker_Unicom_6,
                             'gNB_cite_from_Unicom_12': gNB_cite_from_Unicom_12})
# 只取“小计”行
a = df_5G.loc[df_5G['makers_6'].isin(['小计'])].index
b = df_5G_Unicom.loc[df_5G_Unicom['maker_Unicom_6'].isin(['小计', '合计'])].index
df_5G_total = df_5G.loc[a]
df_5G_Unicom_total = df_5G_Unicom.loc[b]

# 排序
df_5G_total.index = df_5G_total.loc[:, 'city']
df_5G_Unicom_total.index = df_5G_Unicom_total.loc[:, 'city']
df_5G_total.index = df_5G_total.index.astype(custom_order)
df_5G_Unicom_total.index = df_5G_Unicom_total.index.astype(custom_order)
df_5G_total = df_5G_total.sort_index()
df_5G_Unicom_total = df_5G_Unicom_total.sort_index()

# 写入周报
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=106, max_row=117):
    for cell in col:
        value = df_5G_total.at[cell.row - 106, 'CQI_5G_118']
        cell.number_format = '0.00%'
        cell.value = value
for col in week_ws.iter_cols(min_col=3, max_col=3, min_row=106, max_row=117):
    for cell in col:
        value = df_5G_total.at[cell.row - 106, 'SgNB_add_127']
        cell.number_format = '0.00%'
        cell.value = value
for col in week_ws.iter_cols(min_col=4, max_col=4, min_row=106, max_row=117):
    for cell in col:
        value = df_5G_total.at[cell.row - 106, 'SgNB_change_128']
        cell.number_format = '0.00%'
        cell.value = value
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=142, max_row=153):
    for cell in col:
        value = df_5G_total.at[cell.row - 142, 'gNB_cite_total_8']
        cell.number_format = '0'
        cell.value = value
for col in week_ws.iter_cols(min_col=3, max_col=3, min_row=142, max_row=153):
    for cell in col:
        value = df_5G_total.at[cell.row - 142, 'gNB_cite_share_to_Unicom_12']
        cell.number_format = '0'
        cell.value = value
for col in week_ws.iter_cols(min_col=4, max_col=4, min_row=142, max_row=153):
    for cell in col:
        value = df_5G_Unicom_total.at[cell.row - 142, 'gNB_cite_from_Unicom_12']
        cell.number_format = '0'
        cell.value = value
week_wb.save(file_out)

# 6, 1X话务量
df_1x_Erlang = pd.read_excel(file_1X_Erlang)
# 排序
df_1x_Erlang.index = df_1x_Erlang.loc[:, '地市']
df_1x_Erlang.index = df_1x_Erlang.index.astype(custom_order)
df_1x_Erlang = df_1x_Erlang.sort_index()
# 写入周报
for col in week_ws.iter_cols(min_col=2, max_col=2, min_row=3, max_row=13):
    for cell in col:
        value = df_1x_Erlang.at[cell.row - 3, '地市日均话务量(万Erl)-不含切换']
        cell.number_format = '0.00'
        cell.value = value
week_wb.save(file_out)

# 7, RSRP-110dBm覆盖率
df_rsrp = pd.read_excel(file_rsrp, sheet_name='Sheet0')
df_rsrp = RankByCity.rank(df_rsrp, '地市', ['浙江'])
# 写入周报
for col in week_ws.iter_cols(min_col=3, max_col=3, min_row=124, max_row=135):
    for cell in col:
        value = df_rsrp.at[cell.row - 124, '覆盖率']
        cell.number_format = '0.00%'
        cell.value = value/100
week_wb.save(file_out)
print('Congratulations!')
