import pandas as pd
import numpy as np
from pandas.api.types import CategoricalDtype
import openpyxl
from rank_by_city import RankByCity
from complaindata.month_report import MonthReport


file_5G_user = 'E:/投诉分析/202101/202102分析表/5G活跃用户数日报.xlsx'
file_5G_complain = 'E:/投诉分析/202101/202102分析表/咨询投诉清单宽表5G.xlsx'
file_out = 'E:/投诉分析/202101/202102分析表/5G投诉.xlsx'

#
city = ['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省']
# 排序标准序列
custom_order = CategoricalDtype(['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省'], ordered=True)

wb_5G = openpyxl.load_workbook(file_out)
ws_5G = wb_5G['Sheet1']

# 投诉量
complain_5G = MonthReport(file_5G_complain)
df_complain = complain_5G.__calculate__('地市', 'sheet1')
df_complain = df_complain.drop('其他')
df_complain.index = df_complain.index.astype(custom_order)
# print(df_complain.index)
# print(df_complain)
for col in ws_5G.iter_cols(min_col=2, max_col=2, min_row=2, max_row=13):
    for cell in col:
        value = df_complain.at[cell.row - 2, '投诉单']
        cell.value = value
wb_5G.save(file_out)

# 投诉用户
df_user = pd.read_excel(file_5G_user)
df_user = df_user.loc[:, ['分公司', '5G月累计数据活跃用户数']].fillna('全省分公司')
df_user['分公司'] = list(map(lambda x: x[-4::-1][::-1], df_user['分公司']))
df_user = RankByCity.rank(df_user, '分公司')
# print(df_user)
for col in ws_5G.iter_cols(min_col=7, max_col=7, min_row=2, max_row=13):
    for cell in col:
        value = df_user.at[cell.row - 2, '5G月累计数据活跃用户数']
        cell.value = value
wb_5G.save(file_out)

# 画图的表
for col in ws_5G.iter_cols(min_col=1, max_col=1, min_row=18, max_row=28):
    for cell in col:
        cell.value = df_user.at[cell.row - 18, '分公司']

for col in ws_5G.iter_cols(min_col=2, max_col=2, min_row=18, max_row=28):
    for cell in col:
        value = df_complain.at[cell.row - 18, '投诉单'] * 10000 / df_user.at[cell.row - 18, '5G月累计数据活跃用户数']
        cell.number_format = '0.00'
        cell.value = value
wb_5G.save(file_out)


# 投诉原因
df_reason = pd.read_excel(file_5G_complain, 'sheet1')
df_reason = pd.DataFrame(df_reason['原因四层'].value_counts())
print(df_reason)
