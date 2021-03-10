import pandas as pd
from pandas import ExcelFile
from pandas.api.types import CategoricalDtype
import openpyxl
from openpyxl import worksheet

file_NB = 'C:/Users/shelly/Desktop/月度数据源/NB关键指标报表.xls'
file_LTE = 'C:/Users/shelly/Desktop/月度数据源/LTE关键指标统计报表--lte性能-月度.xls'
file_out = 'C:/Users/shelly/Desktop/月度数据源/网优中心移动网运行月报数据-12月份).xlsx'
# 读取源数据并排序


# df2.sort_values(by='市/地区/州/盟')
# 先看能否写入
wb = openpyxl.load_workbook(file_out)
ws = wb['sheet1']
# print(type(ws))
# print('模板已导入')

city = ['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '浙江']
data = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
df1 = pd.DataFrame({"city": city, "data": data})
# print(df1.info())
# print(df1)
base = []
rrc = []
# 从1开始计数
for col in ws.iter_cols(min_col=1, max_col=1, min_row=4, max_row=15):
    for cell in col:
        value = df1.at[cell.row-4, 'data']
        # cell.value = value
        # excel 数据填入数组
        base.append(cell.value)

for col in ws.iter_cols(min_col=3, max_col=3, min_row=4, max_row=15):
    for cell in col:
        value = df1.at[cell.row-4, 'data']
        # cell.value = value
        # excel 数据填入数组
        rrc.append(cell.value)
df2 = pd.DataFrame({"城市": base, "rrc连接": rrc})
print(df2)
df2.index = df2.iloc[:, 0]
print(df2)
# df3 = df2.sort_values(by=["rrc连接"])
print("自定义排序")
custom_order = CategoricalDtype(['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省'], ordered=True)
df2.index = df2.index.astype(custom_order)
final = df2.sort_index()
print(final)
#wb.save(file_out)

# 筛选
# 此行是注释df[np.logical_and(df['one']> 5,df['two']>5)] 或者 df[(df['one']> 5) & (df['two']>5)]
# 第一步多列筛选；第二步指定列进行字符频率统计
# selection = data2.loc[(data2['支付方式'].isin(['支付宝付款码', '微信付款码'])) & (data2['一级类目'].isin(['环球美食', '个人洗护', '美容彩妆']))]. \
                # loc[:, '一级类目'].value_counts()