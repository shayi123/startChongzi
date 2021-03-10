import pandas as pd
from pandas.api.types import CategoricalDtype
import openpyxl
from rank_by_city import RankByCity
from openpyxl import worksheet

# 先将文件保存为.xlsx格式。因为openpyxl不支持.xls文件
file_out = 'C:/Users/shelly/Desktop/月度数据源/网优中心移动网运行月报数据-11月份.xlsx'
file_NB = 'C:/Users/shelly/Desktop/月度数据源/NB关键指标报表.xlsx'
file_LTE_major = 'C:/Users/shelly/Desktop/月度数据源/LTE关键指标统计报表-major性能-月度.xlsx'
file_lte_dpi = 'C:/Users/shelly/Desktop/月度数据源/lte_dpi.xlsx'
file_weak_cell = 'C:/Users/shelly/Desktop/月度数据源/问题小区统计.xlsx'
file_volte_conn = 'C:/Users/shelly/Desktop/月度数据源/volte网络连通率.xlsx'
file_lte_MR = 'C:/Users/shelly/Desktop/月度数据源/MR覆盖.xlsx'
# RRC连接成功率 volte信令建立成功率
file_rrc_volte_succ = 'C:/Users/shelly/Desktop/月度数据源/rrc volte 连接成功率_000.xlsx'
file_volte_major = 'C:/Users/shelly/Desktop/月度数据源/VOLTE关键指标报表-月.xlsx'
# 小区退服时长统计.xlsx
file_crash_duration = 'C:/Users/shelly/Desktop/月度数据源/小区退服时长统计.xlsx'

#
city_default = ['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省']
# 排序标准序列
custom_order = CategoricalDtype(['杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州', '丽水', '台州', '舟山', '全省'], ordered=True)

# 获取评价总表
evaluation_wb = openpyxl.load_workbook(file_out)
evaluation_sheet = evaluation_wb['评价表']
# testws = evaluation_wb['覆盖优']
# print(type(testws))
# for col in testws.iter_cols(min_col=1, max_col=1, min_row=3, max_row=5):
#     print('ok')
# 1, NB
NB_wb = openpyxl.load_workbook(file_NB)
NB_ws = NB_wb['sheet1']

city = []
NB_rrc = []
# 读取源数据并排序
for col in NB_ws.iter_cols(min_col=2, max_col=2, min_row=4, max_row=15):
    for cell in col:
        city.append(cell.value)
for col in NB_ws.iter_cols(min_col=5, max_col=5, min_row=4, max_row=15):
    for cell in col:
        NB_rrc.append(cell.value)
df_NB = pd.DataFrame({"城市": city, "rrc连接成功率": NB_rrc})
# print(df_NB)
# '浙江'都换作'全省'。 先得到元素所在坐标，再赋值
a = df_NB.loc[df_NB['城市'].isin(['浙江'])].index.tolist()
df_NB.loc[a[0], '城市'] = '全省'
# print(df_NB)
# 指定df的index为city
df_NB.index = df_NB.iloc[:, 0]
# 排序
df_NB.index = df_NB.index.astype(custom_order)
df_NB = df_NB.sort_index()
# print(df_NB)

# 将NB数据写入评价表
for col in evaluation_sheet.iter_cols(min_col=27, max_col=27, min_row=4, max_row=15):
    for cell in col:
        value = df_NB.at[cell.row - 4, 'rrc连接成功率']
        # 设置格式为数值，保留两位小数
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)

# 2, LTE-DPI感知优良率

# 读取源数据并排序
data_dpi = pd.read_excel(file_lte_dpi, sheet_name='kqi')
dpi_avg = data_dpi.groupby('省市')['总体优良率'].mean()
# print(type(dpi_avg))
# print(data_dpi)
lte_dpi = pd.DataFrame(dpi_avg)
# print(type(lte_dpi))
# print(lte_dpi.index)
lte_dpi.index = lte_dpi.index.astype(custom_order)
lte_dpi = lte_dpi.sort_index()
# print(type(lte_dpi))
# print(lte_dpi)
# 将lte_dpi数据写入评价表
for col in evaluation_sheet.iter_cols(min_col=2, max_col=2, min_row=4, max_row=15):
    for cell in col:
        value = lte_dpi.at[cell.row - 4, '总体优良率']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
evaluation_wb.save(file_out)

# 3, LTE 大部分指标 file_LTE_major
lte_major_wb = openpyxl.load_workbook(file_LTE_major)
lte_major_ws = lte_major_wb['sheet1']

city = []
# 4G下切3G比例
G4_to_G3_5 = []
# 用户下行调度吞吐率
DL_rate_6 = []
# CQI优良比
CQI_8 = []
# 无线连接成功率
wireless_success_9 = []
# RRC连接重建比例
rrc_rebuild_ratio_11 = []
# E-RAB掉线率
ERAB_failure_ratio_13 = []
# 系统内切换成功率
inner_switch_14 = []

for col in lte_major_ws.iter_cols(min_col=2, max_col=2, min_row=5, max_row=16):
    for cell in col:
        city.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=12, max_col=12, min_row=5, max_row=16):
    for cell in col:
        G4_to_G3_5.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=7, max_col=7, min_row=5, max_row=16):
    for cell in col:
        DL_rate_6.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=13, max_col=13, min_row=5, max_row=16):
    for cell in col:
        CQI_8.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=9, max_col=9, min_row=5, max_row=16):
    for cell in col:
        wireless_success_9.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=10, max_col=10, min_row=5, max_row=16):
    for cell in col:
        rrc_rebuild_ratio_11.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=11, max_col=11, min_row=5, max_row=16):
    for cell in col:
        ERAB_failure_ratio_13.append(cell.value)
for col in lte_major_ws.iter_cols(min_col=8, max_col=8, min_row=5, max_row=16):
    for cell in col:
        inner_switch_14.append(cell.value)
df_lte_major = pd.DataFrame({'city': city, '4G下切3G比例(%)': G4_to_G3_5, '用户下行调度吞吐率(QCI9) (Mbps)': DL_rate_6,
                             'CQI优良比(%)': CQI_8, '无线连接成功率 (%)': wireless_success_9,
                             'RRC连接重建比例 (%)': rrc_rebuild_ratio_11, 'E-RAB掉线率 (%)': ERAB_failure_ratio_13,
                             '系统内切换成功率 (%)': inner_switch_14})
# 排序
df_lte_major = RankByCity.rank(df_lte_major, 'city', ['浙江'])
# print(df_lte_major)
# 写入月报
for col in evaluation_sheet.iter_cols(min_col=5, max_col=5, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, '4G下切3G比例(%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=6, max_col=6, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, '用户下行调度吞吐率(QCI9) (Mbps)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=8, max_col=8, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, 'CQI优良比(%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=9, max_col=9, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, '无线连接成功率 (%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=11, max_col=11, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, 'RRC连接重建比例 (%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=13, max_col=13, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, 'E-RAB掉线率 (%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=14, max_col=14, min_row=4, max_row=15):
    for cell in col:
        value = df_lte_major.at[cell.row - 4, '系统内切换成功率 (%)']
        # 设置格式为数值，保留3位小数
        cell.number_format = '0.000'
        cell.value = value
evaluation_wb.save(file_out)

# 4, RRC重建成功率
rrc_rebuild_success_10 = []
# VOLTE 信令建立成功率
signal_build_ratio_20 = []

rrc_volte_succ = pd.read_excel(file_rrc_volte_succ)
rrc_volte_succ = rrc_volte_succ.loc[rrc_volte_succ['厂家'] == '不区分厂家']
rrc_volte_succ = RankByCity.rank(rrc_volte_succ, '地市')
# print(rrc_volte_succ.loc[:, ['地市', '3.29 RRC连接重建成功率(%)', '2.20 E-RAB建立成功率(QCI5)(%)']])
# 写入评价表 省数据在另一个excel中 TODO
for col in evaluation_sheet.iter_cols(min_col=10, max_col=10, min_row=4, max_row=14):
    for cell in col:
        value = rrc_volte_succ.at[cell.row - 4, '3.29 RRC连接重建成功率(%)']
        cell.number_format = '0.00'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=20, max_col=20, min_row=4, max_row=14):
    for cell in col:
        value = rrc_volte_succ.at[cell.row - 4, '2.20 E-RAB建立成功率(QCI5)(%)']
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)

# 5, 问题小区 weak_cell
# 重叠覆盖
overlay_15 = []
# 模3干扰
mod3_16 = []
# 越区覆盖
outbounds_cell_17 = []

weak_cell = pd.read_excel(file_weak_cell)
overlay = weak_cell[weak_cell['问题分类'] == '重叠覆盖'].loc[:, '市/区'].value_counts()
mod3 = weak_cell[weak_cell['问题分类'] == 'MOD3干扰'].loc[:, '市/区'].value_counts()
outbounds_cell = weak_cell[weak_cell['问题分类'] == '过覆盖'].loc[:, '市/区'].value_counts()

value_overlay = [0] * 12
value_mods3 = [0] * 12
value_outbounds_cell = [0] * 12
df_weak_cell = pd.DataFrame({'city': city_default, '重叠覆盖': value_overlay, 'MOD3干扰': value_mods3, '越区覆盖': value_outbounds_cell})

for i, v in overlay.items():
    df_weak_cell.loc[df_weak_cell.loc[df_weak_cell['city'].isin([i])].index.tolist()[0], '重叠覆盖'] = v
    df_weak_cell.loc[11, '重叠覆盖'] += v
for i, v in mod3.items():
    df_weak_cell.loc[df_weak_cell.loc[df_weak_cell['city'].isin([i])].index.tolist()[0], 'MOD3干扰'] = v
    df_weak_cell.loc[11, 'MOD3干扰'] += v
for i, v in outbounds_cell.items():
    df_weak_cell.loc[df_weak_cell.loc[df_weak_cell['city'].isin([i])].index.tolist()[0], '越区覆盖'] = v
    df_weak_cell.loc[11, '越区覆盖'] += v
df_weak_cell = RankByCity.rank(df_weak_cell, 'city')
# print(df_weak_cell)
# 写入评价表
for col in evaluation_sheet.iter_cols(min_col=15, max_col=15, min_row=4, max_row=15):
    for cell in col:
        value = df_weak_cell.at[cell.row - 4, '重叠覆盖']
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=16, max_col=16, min_row=4, max_row=15):
    for cell in col:
        value = df_weak_cell.at[cell.row - 4, 'MOD3干扰']
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=17, max_col=17, min_row=4, max_row=15):
    for cell in col:
        value = df_weak_cell.at[cell.row - 4, '越区覆盖']
        cell.value = value
evaluation_wb.save(file_out)

# 6, 退服时长
city_crash_cell = []
crash_duration_18 = []

crash_cell_wb = openpyxl.load_workbook(file_crash_duration)
crash_cell_ws = crash_cell_wb['sheet1']

for col in crash_cell_ws.iter_cols(min_col=1, max_col=1, min_row=4, max_row=15):
    for cell in col:
        city_crash_cell.append(cell.value)
for col in crash_cell_ws.iter_cols(min_col=7, max_col=7, min_row=4, max_row=15):
    for cell in col:
        crash_duration_18.append(cell.value)

df_crash_duration = pd.DataFrame({'city': city_crash_cell, '退服时长': crash_duration_18})
df_crash_duration = RankByCity.rank(df_crash_duration, 'city', ['浙江'])
# print(df_crash_duration)
# 写入评价表
for col in evaluation_sheet.iter_cols(min_col=18, max_col=18, min_row=4, max_row=15):
    for cell in col:
        value = df_crash_duration.at[cell.row - 4, '退服时长']
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)

# 7, VOLTE 大部分指标 file_VOLTE_major
city_volte_major = []
# 语音连通率
video_connect_ratio_19 = []
# VoLTE语音上行丢包率 (%)
video_ul_missing_22 = []
# VoLTE语音下行丢包率 (%)
video_dl_missing_23 = []
# 语音掉话率
video_lose_ratio_24 = []
# VoLTE信令掉线率 (%)
signal_lose_ratio_25 = []


volte_major_wb = openpyxl.load_workbook(file_volte_major)
volte_major_ws = volte_major_wb['sheet1']

for col in volte_major_ws.iter_cols(min_col=2, max_col=2, min_row=4, max_row=15):
    for cell in col:
        city_volte_major.append(cell.value)
for col in volte_major_ws.iter_cols(min_col=5, max_col=5, min_row=4, max_row=15):
    for cell in col:
        video_connect_ratio_19.append(cell.value)
for col in volte_major_ws.iter_cols(min_col=6, max_col=6, min_row=4, max_row=15):
    for cell in col:
        video_lose_ratio_24.append(cell.value)
for col in volte_major_ws.iter_cols(min_col=7, max_col=7, min_row=4, max_row=15):
    for cell in col:
        video_ul_missing_22.append(cell.value)
for col in volte_major_ws.iter_cols(min_col=8, max_col=8, min_row=4, max_row=15):
    for cell in col:
        video_dl_missing_23.append(cell.value)
for col in volte_major_ws.iter_cols(min_col=9, max_col=9, min_row=4, max_row=15):
    for cell in col:
        signal_lose_ratio_25.append(cell.value)

# 构造df
df_volte_major = pd.DataFrame({'city': city_volte_major, '语音连通率': video_connect_ratio_19, 'VoLTE语音上行丢包率': video_ul_missing_22,
                               'VOLTE语音下行丢包率': video_dl_missing_23, '语音掉话率': video_lose_ratio_24, 'VoLTE信令掉线率': signal_lose_ratio_25})
# 排序
df_volte_major = RankByCity.rank(df_volte_major, 'city', ['浙江'])
# print(df_volte_major.index)
# 写入评价表
for col in evaluation_sheet.iter_cols(min_col=19, max_col=19, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_major.at[cell.row - 4, '语音连通率']
        cell.number_format = '0.00'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=22, max_col=22, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_major.at[cell.row - 4, 'VoLTE语音上行丢包率']
        cell.number_format = '0.00'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=23, max_col=23, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_major.at[cell.row - 4, 'VOLTE语音下行丢包率']
        cell.number_format = '0.00'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=24, max_col=24, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_major.at[cell.row - 4, '语音掉话率']
        cell.number_format = '0.00'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=25, max_col=25, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_major.at[cell.row - 4, 'VoLTE信令掉线率']
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)

# 8, VOLTE网络连通率 呼叫时延
network_connect_21 = []
calling_delay_26 = []

volte_conn = pd.read_excel(file_volte_conn)
df_volte_conn = pd.DataFrame(volte_conn.groupby('省/地市')['volte网络接通率'].mean())
df_calling_delay = pd.DataFrame(volte_conn.groupby('省/地市')['呼叫建立时延'].mean() / 1000)
# 排序
df_calling_delay.index = df_calling_delay.index.astype(custom_order)
df_volte_conn.index = df_volte_conn.index.astype(custom_order)
df_volte_conn = df_volte_conn.sort_index()
df_calling_delay = df_calling_delay.sort_index()
# print(df_volte_conn)
# print(df_calling_delay)
# 写入评价表
for col in evaluation_sheet.iter_cols(min_col=21, max_col=21, min_row=4, max_row=15):
    for cell in col:
        value = df_volte_conn.at[cell.row - 4, 'volte网络接通率']
        cell.number_format = '0.000'
        cell.value = value
for col in evaluation_sheet.iter_cols(min_col=26, max_col=26, min_row=4, max_row=15):
    for cell in col:
        value = df_calling_delay.at[cell.row - 4, '呼叫建立时延']
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)

# 9, MR覆盖率
city_MR = []
band = []
RSRP = []

MR_wb = openpyxl.load_workbook(file_lte_MR)
MR_ws = MR_wb['cover_lan_list']

for col in MR_ws.iter_cols(min_col=4, max_col=4, min_row=3):
    for cell in col:
        city_MR.append(cell.value)
for col in MR_ws.iter_cols(min_col=5, max_col=5, min_row=3):
    for cell in col:
        band.append(cell.value)
for col in MR_ws.iter_cols(min_col=13, max_col=13, min_row=3):
    for cell in col:
        RSRP.append(cell.value)

# 构造df
df_RSRP = pd.DataFrame({'city': city_MR, '频段': band, 'RSRP覆盖率': RSRP})
df_RSRP = df_RSRP.loc[df_RSRP['频段'] == '汇总']
# 排序
df_RSRP = RankByCity.rank(df_RSRP, 'city')
# print(df_RSRP)
# 写入评价表
for col in evaluation_sheet.iter_cols(min_col=3, max_col=3, min_row=4, max_row=15):
    for cell in col:
        value = df_RSRP.at[cell.row - 4, 'RSRP覆盖率']
        cell.number_format = '0.00'
        cell.value = value
evaluation_wb.save(file_out)
print('Congratulations!')