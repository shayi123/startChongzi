# 今天开始学习数据分析，从numpy和pandas学起

# 读取文件
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import openpyxl

# numpy 读取csv
data = np.genfromtxt('/Users/shelly/PycharmProjects/startChongzi/numpylearn/sales.csv', dtype=float, delimiter=',', skip_header=1, usecols=(6, 7), encoding='utf-8')
print(data)
print(data.size)

# openpyxl 读取 xlsx
sales = load_workbook('/Users/shelly/PycharmProjects/startChongzi/numpylearn/sales.xlsx')
sheet_names = sales.sheetnames
sheet1 = sales[sheet_names[0]]
cell = sheet1.cell(1, 1).value
print(cell)

# openpyxl 创建xlsx并保存，同名文件会覆盖
outwb = openpyxl.Workbook()
outsheet = outwb.create_sheet(index=0)
save_path = '/Users/shelly/PycharmProjects/startChongzi/numpylearn/out.xlsx'
outwb.save(save_path)

# pandas

# pd.ExcelFile()
data2 = pd.read_excel('/Users/shelly/PycharmProjects/startChongzi/numpylearn/sales.xlsx')
print(data2.size)
print(data2)


